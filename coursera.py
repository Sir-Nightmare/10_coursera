import argparse
import requests
import json

from lxml import etree
from openpyxl import Workbook
from bs4 import BeautifulSoup
from os.path import join
from random import sample

COURSES_XML_URL = 'https://www.coursera.org/sitemap~www~courses.xml'


def convert_soup_to_text(tag):
    return tag.text if tag else None


def get_courses_urls(xml_url, quantity=20):
    xml_page = requests.get(xml_url)
    root = etree.fromstring(xml_page.content)
    urls = [url.text for url in root.iter('{*}loc')]
    return sample(urls, quantity)


def get_datetime_course(soup):
    json_course = convert_soup_to_text(soup.find('script', {'type': 'application/ld+json'}))
    if json_course and 'startDate' in json_course:
        return json.loads(json_course)['hasCourseInstance'][0]['startDate']


def get_course_info(course_url):
    page = requests.get(course_url).content
    soup = BeautifulSoup(page, 'html.parser')
    course_name = soup.find('div', {'class': 'title display-3-text'}).text
    course_lang = soup.find('div', {'class': 'language-info'}).text
    course_date = get_datetime_course(soup)
    duration = len(soup.find_all('div', {'class': 'week'}))
    average_score = convert_soup_to_text(soup.find('div', {'class': 'ratings-text bt3-visible-xs'}))
    course_info = {
        'Title': course_name,
        'Language': course_lang,
        'Start Date': course_date,
        'Duration (weeks)': duration,
        'Course Rate': average_score,
        'URL': course_url
    }
    return course_info


def output_info_to_workbook(filepath, courses_info):
    headers = ['Title', 'Language', 'Start Date', 'Duration (weeks)', 'Course Rate', 'URL']
    wb = Workbook()
    sheet = wb.active
    for num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=num).value = header
    for num, info_about_course in enumerate(courses_info, 2):
        for i, key in enumerate(headers, 1):
            sheet.cell(row=num, column=i).value = info_about_course[key]
    return wb


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-p', '--path', default='', type=str, help='Input full path to folder')
    parser.add_argument('-n', '--number', default=20, type=int, help='Input number of courses')
    args = parser.parse_args()
    links = get_courses_urls(COURSES_XML_URL, args.number)
    print('Getting courses info has been started')
    courses_info = [get_course_info(link) for link in links]
    work_book = output_info_to_workbook(args.path, courses_info)
    work_book.save(join(args.path, 'Courses from Coursera.xlsx'))
    print('Done!')
